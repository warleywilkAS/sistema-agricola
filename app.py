import pandas as pd
from io import BytesIO
from flask import send_file
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify
from models import db, FormularioSoja, Pulverizacao
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///database.db'
app.config['SECRET_KEY'] = 'chave-secreta-para-formulario-agricola-2026'
db.init_app(app)

# Listas de alvos
INSETOS_ALVO = [
    "Lagarta da soja (Anticarsia gemmatalis)",
    "Lagarta das vagens (Spodoptera spp.)",
    "Lagarta falsa medideira (Chrysodeixis includens)",
    "Lagartas do grupo Heliothinae",
    "Percevejo barriga verde (Dichelops spp.)",
    "Percevejo marrom (Euschistus heros)",
    "Percevejo verde (Nezara viridula)",
    "Percevejo verde pequeno (Piezodorus guildinii)",
    "Broca dos ponteiros (Crocidosema aporema)",
    "Mosca Branca",
    "Outros insetos praga",
    "Tamanduá da soja (Sternechus subsignatus)",
    "Tripes",
    "Vaquinhas (Diabrotica/ Cerotoma/ Colapsis)"
]

DOENCAS_ALVO = [
    "Antracnose (Colletotrichum truncatum)",
    "Cancro da haste (Diaporthe spp.)",
    "Ferrugem asiática (Phakopsora pachyrhizi)",
    "Mancha alvo (Corynespora cassicola)",
    "Mancha de cercospora (Cercospora kikuchii)",
    "Mancha olho-de-rã (Cercospora sojina)",
    "Mancha parda (Septoria glycines)",
    "Mela ou requeima (Rhizoctonia solani)",
    "Mofo branco (Sclerotinia sclerotiorum)",
    "Mildio (Peronospora manshurica)",
    "Oídio (Microsphaera diffusa)",
    "Outras Doenças Fungicas"
]

PLANTAS_DANINHAS = [
    "Buva (Conyza spp.)",
    "Capim-amargoso (Digitaria insularis)",
    "Caruru (Amaranthus spp.)",
    "Capim-pé-de-galinha (Eleusine indica)",
    "Leiteiro (Euphorbia heterophylla)",
    "Picão-preto (Bidens pilosa)",
    "Trapoeraba (Commelina spp.)",
    "Outras Plantas Daninhas"
]

ACAROS = [
    "Ácaro-rajado (Tetranychus urticae)",
    "Ácaro-verde (Mononychellus planki)",
    "Ácaro-branco (Polyphagotarsonemus latus)",
    "Ácaros-vermelhos (Tetranychus spp.)",
    "Outros ácaros"
]

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/form', methods=['GET', 'POST'])
def form():
    if request.method == 'POST':
        try:
            # Criar novo formulário
            formulario = FormularioSoja()
            
            # Identificação
            formulario.numero_produtor = request.form.get('numero_produtor')
            formulario.regiao = request.form.get('regiao')
            formulario.municipio = request.form.get('municipio')
            formulario.meso_idr = request.form.get('meso_idr')
            formulario.area_soja = float(request.form.get('area_soja') or 0)
            formulario.produtividade_media = float(request.form.get('produtividade_media') or 0)
            formulario.cultivar = request.form.get('cultivar')
            formulario.bt = request.form.get('bt')
            formulario.data_plantio = request.form.get('data_plantio')
            formulario.data_emergencia = request.form.get('data_emergencia')
            formulario.houve_adversidade = request.form.get('houve_adversidade')
            formulario.qual_adversidade = request.form.get('qual_adversidade')
            formulario.nome_coletor = request.form.get('nome_coletor')
            formulario.unidade_municipal = request.form.get('unidade_municipal')
            
            # Conhecimento MIP e MID
            formulario.conhecimento_mid = request.form.get('conhecimento_mid')
            formulario.utiliza_mid = request.form.get('utiliza_mid')
            formulario.conhecimento_mip = request.form.get('conhecimento_mip')
            formulario.utiliza_mip = request.form.get('utiliza_mip')
            
            # Controle Plantas Invasoras
            formulario.herbicida_dessecacao_alvo = request.form.get('herbicida_dessecacao_alvo')
            formulario.herbicida_dessecacao_aplicacoes = int(request.form.get('herbicida_dessecacao_aplicacoes') or 0)
            formulario.herbicida_pre_alvo = request.form.get('herbicida_pre_alvo')
            formulario.herbicida_pre_aplicacoes = int(request.form.get('herbicida_pre_aplicacoes') or 0)
            formulario.herbicida_pos_alvo = request.form.get('herbicida_pos_alvo')
            formulario.herbicida_pos_aplicacoes = int(request.form.get('herbicida_pos_aplicacoes') or 0)
            
            # Outras informações
            formulario.tratamento_sementes = request.form.get('tratamento_sementes')
            formulario.sal_mistura = request.form.get('sal_mistura')
            formulario.controle_biologico = request.form.get('controle_biologico')
            
            # FBN
            formulario.inoculacao_sementes = request.form.get('inoculacao_sementes')
            formulario.forma_inoculacao = request.form.get('forma_inoculacao')
            formulario.coinoculacao = request.form.get('coinoculacao')
            formulario.co_mo = request.form.get('co_mo')
            formulario.co_mo_aplicacao = request.form.get('co_mo_aplicacao')
            
            db.session.add(formulario)
            db.session.flush()  # Para obter o ID
            
                        # Salvar pulverizações
            # Pré-plantio com múltiplas classes
            if request.form.get('data_pre_plantio'):  # <--- AGORA COM INDENTAÇÃO CORRETA!
                classes_pre = request.form.getlist('classe_pre_plantio')
                if classes_pre:
                    classe_pre_str = ', '.join(classes_pre)
                else:
                    classe_pre_str = ''
                
                alvo_pre = request.form.get('alvo_pre_plantio')
                
                if classe_pre_str and alvo_pre:
                    pulv = Pulverizacao(
                        formulario_id=formulario.id,
                        tipo='pre_plantio',
                        data=request.form.get('data_pre_plantio'),
                        classe_produto=classe_pre_str,
                        alvo=alvo_pre
                    )
                    db.session.add(pulv)
            
            # Pulverizações pós-emergência (até 7)
            for i in range(1, 8):
                data = request.form.get(f'data_pos_{i}')
                if data:
                    classes = request.form.getlist(f'classe_pos_{i}')
                    if classes:
                        classe_str = ', '.join(classes)
                    else:
                        classe_str = ''
                    
                    alvo = request.form.get(f'alvo_pos_{i}')
                    
                    if classe_str and alvo:
                        pulv = Pulverizacao(
                            formulario_id=formulario.id,
                            tipo=f'pos_{i}',
                            data=data,
                            classe_produto=classe_str,
                            alvo=alvo
                        )
                        db.session.add(pulv)
            
            db.session.commit()
            flash('Formulário salvo com sucesso!', 'success')
            return redirect(url_for('view_records'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao salvar: {str(e)}', 'danger')
            return redirect(url_for('form'))
    
    return render_template('form.html', 
                      insetos=INSETOS_ALVO, 
                      doencas=DOENCAS_ALVO,
                      plantas=PLANTAS_DANINHAS,
                      acaros=ACAROS)

@app.route('/record/<int:id>')
def view_record(id):
    registro = FormularioSoja.query.get_or_404(id)
    return render_template('view_record.html', 
                          registro=registro, 
                          insetos=INSETOS_ALVO, 
                          doencas=DOENCAS_ALVO,
                          plantas=PLANTAS_DANINHAS,
                          acaros=ACAROS)

@app.route('/records')
def view_records():
    registros = FormularioSoja.query.order_by(FormularioSoja.data_criacao.desc()).all()
    return render_template('view_records.html', registros=registros)

@app.route('/export/excel')
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from datetime import datetime
    
    # Buscar todos os registros
    registros = FormularioSoja.query.order_by(FormularioSoja.data_criacao.desc()).all()
    
    # Criar workbook
    wb = Workbook()
    
    # ============================================================
    # ABA 1: BD (Base de Dados - Dicionário)
    # ============================================================
    ws_bd = wb.active
    ws_bd.title = "BD"
    
    # Cabeçalhos da aba BD (baseado na planilha modelo)
    cabecalhos_bd = [
        "Doenças", "Bactérias", "Pragas", "Ácaros", "", "", "Estádio", "Outros", 
        "Aplicações", "Evento", "", "Cultivares", "", "", "", "", "MACROS", "",
        "Regionais", "Apucarana", "Campo_Mourão", "Cascavel", "Cianorte", 
        "Cornélio_Procópio", "Curitiba", "Dois_Vizinhos", "Francisco_Beltrão",
        "Guarapuava", "Irati", "Ivaiporã", "Laranjeiras_do_Sul", "Londrina",
        "Maringá", "Paranaguá", "Paranavaí", "Pato_Branco", "Ponta_Grossa",
        "Sto_Antônio_da_Platina", "Toledo", "Umuarama", "União_da_Vitória", "",
        "NOROESTE", "NORTE", "OESTE", "SUDOESTE", "SUL", "Plantas Invasoras"
    ]
    
    for col, titulo in enumerate(cabecalhos_bd, 1):
        ws_bd.cell(row=1, column=col, value=titulo)
    
    # Preencher com dados das listas
    # Doenças (col A)
    for i, doenca in enumerate(DOENCAS_ALVO, 2):
        ws_bd.cell(row=i, column=1, value=doenca)
    
    # Pragas (col C)
    for i, praga in enumerate(INSETOS_ALVO, 2):
        ws_bd.cell(row=i, column=3, value=praga)
    
    # Ácaros (col D)
    for i, acaro in enumerate(ACAROS, 2):
        ws_bd.cell(row=i, column=4, value=acaro)
    
    # Plantas invasoras (col BH)
    for i, planta in enumerate(PLANTAS_DANINHAS, 2):
        ws_bd.cell(row=i, column=60, value=planta)
    
    # Estádios fenológicos (col G)
    estadios = ["VE", "VC", "V1", "V2", "V3", "V4", "V5", "V6", "V7", "V8", "V9", "VN",
                "R1", "R2", "R3", "R4", "R5.1", "R5.2", "R5.3", "R5.4", "R5.5",
                "R6", "R7.1", "R7.2", "R7.3", "R7.4", "R8.1", "R8.2"]
    for i, est in enumerate(estadios, 2):
        ws_bd.cell(row=i, column=7, value=est)
    
    # ============================================================
    # ABA 2: Total_Pr (Dados Consolidados)
    # ============================================================
    ws_total = wb.create_sheet("Total_Pr")
    
    # Configurar estilos
    titulo_fill = PatternFill(start_color="2c5f2d", end_color="2c5f2d", fill_type="solid")
    titulo_font = Font(color="FFFFFF", bold=True)
    sub_fill = PatternFill(start_color="e9ecef", end_color="e9ecef", fill_type="solid")
    
    # LINHA 4: Títulos principais (mesclados)
    titulos_principais = [
        (1, 3, "PLANILHA TABULAÇÃO DADOS QUESTIONÁRIOS APLICAÇÃO DEFENSIVOS PARA CONTROLE PRAGAS E DOENÇAS_PR_SAFRA 19_20_V1"),
        (4, 6, ""),
        (7, 15, "CONHECIMENTO MONITORAMENTO"),
        (16, 19, "3_Informação Plantas Invasoras"),
        (20, 24, "4.0_INFORMAÇÃO _PULVERIZAÇÃO DESSECAÇÃO"),
        (25, 29, "4.1_INFORMAÇÃO _PRIMEIRA PULVERIZAÇÃO APÓS EMERGÊNCIA"),
        (30, 34, "4.2_INFORMAÇÃO _SEGUNDA PULVERIZAÇÃO APÓS EMERGÊNCIA"),
        (35, 39, "4.3_INFORMAÇÃO _TERCEIRA PULVERIZAÇÃO APÓS EMERGÊNCIA"),
        (40, 44, "4.4_INFORMAÇÃO _QUARTA PULVERIZAÇÃO APÓS EMERGÊNCIA"),
        (45, 49, "4.5_INFORMAÇÃO _QUINTA PULVERIZAÇÃO APÓS EMERGÊNCIA"),
        (50, 54, "4.6_INFORMAÇÃO _SEXTA PULVERIZAÇÃO APÓS EMERGÊNCIA"),
        (55, 59, "4.7_INFORMAÇÃO _SÉTIMA PULVERIZAÇÃO APÓS EMERGÊNCIA"),
        (60, 62, "5.OUTRAS INFORMAÇÕES"),
        (63, 63, "6.INOCULAÇÃO"),
    ]
    
    for inicio, fim, titulo in titulos_principais:
        ws_total.merge_cells(start_row=4, start_column=inicio, end_row=4, end_column=fim)
        celula = ws_total.cell(row=4, column=inicio, value=titulo)
        celula.fill = titulo_fill
        celula.font = titulo_font
        celula.alignment = Alignment(horizontal='center', vertical='center')
    
    # LINHA 5: Subtítulos (corrigido!)
    subtitulos = [
        "Tabela", "N° P", "Ordem", "", "Meso_IDR",
        "Região", "Município", "Área com Soja (ha)", "Cultivar",
        "Bt", "Produtividade Média (sc/ha)", "Data Plantio",
        "Adversidade", "Sinistro", "Conhec. MID", "Utiliza MID",
        "Conhec. MIP", "Utiliza MIP", "Classe do Produto",
        "Alvo", "N° Aplicações", "Classe do Produto", "Alvo",
        "N° Aplicações", "Classe do Produto", "Alvo",
        "N° Aplicações", "Classe do Produto", "Alvo",
        "N° Aplicações", "Pulverização na Dessecação", "Data",
        "Classe do Produto", "Alvo_1", "Alvo_2", "Alvo_3",
        # ... continue com todos os subtítulos
    ]

    for col, titulo in enumerate(subtitulos, 1):
        if titulo:  # só preenche se não for string vazia
            celula = ws_total.cell(row=5, column=col, value=titulo)
            celula.fill = sub_fill
    
    # ============================================================
    # POPULAR DADOS
    # ============================================================
    linha_atual = 7
    for reg in registros:
        # Col A: Tabela (ex: TB1, TB2...)
        ws_total.cell(row=linha_atual, column=1, value=f"TB{linha_atual-6}.")
        
        # Col B: N° P
        ws_total.cell(row=linha_atual, column=2, value=linha_atual-6)
        
        # Col C: Ordem (começa com 1 a cada nova tabela)
        ws_total.cell(row=linha_atual, column=3, value=1)
        
        # Col D a G: Meso_IDR, Região, Município, Área
        ws_total.cell(row=linha_atual, column=4, value=reg.meso_idr)
        ws_total.cell(row=linha_atual, column=5, value=reg.regiao)
        ws_total.cell(row=linha_atual, column=6, value=reg.municipio)
        ws_total.cell(row=linha_atual, column=7, value=reg.area_soja)
        
        # Col H: Cultivar
        ws_total.cell(row=linha_atual, column=8, value=reg.cultivar)
        
        # Col I: Bt
        ws_total.cell(row=linha_atual, column=9, value=reg.bt)
        
        # Col J: Produtividade Média
        ws_total.cell(row=linha_atual, column=10, value=reg.produtividade_media)
        
        # Col K: Data Plantio
        ws_total.cell(row=linha_atual, column=11, value=reg.data_plantio)
        
        # Col L: Adversidade
        ws_total.cell(row=linha_atual, column=12, value=reg.qual_adversidade)
        
        # Col M: Sinistro (se houve adversidade)
        ws_total.cell(row=linha_atual, column=13, value="SIM" if reg.houve_adversidade == "SIM" else "")
        
        # Col N-P: Conhecimento MID/MIP
        ws_total.cell(row=linha_atual, column=14, value=reg.conhecimento_mid)
        ws_total.cell(row=linha_atual, column=15, value=reg.utiliza_mid)
        ws_total.cell(row=linha_atual, column=16, value=reg.conhecimento_mip)
        ws_total.cell(row=linha_atual, column=17, value=reg.utiliza_mip)
        
        # Col Q-S: Plantas invasoras
        ws_total.cell(row=linha_atual, column=18, value=reg.herbicida_dessecacao_alvo)
        ws_total.cell(row=linha_atual, column=19, value=reg.herbicida_dessecacao_aplicacoes)
        ws_total.cell(row=linha_atual, column=20, value=reg.herbicida_pre_alvo)
        ws_total.cell(row=linha_atual, column=21, value=reg.herbicida_pre_aplicacoes)
        ws_total.cell(row=linha_atual, column=22, value=reg.herbicida_pos_alvo)
        ws_total.cell(row=linha_atual, column=23, value=reg.herbicida_pos_aplicacoes)
        
        # Col T e seguintes: Pulverizações
        col_pulv = 24
        # Pré-plantio
        for p in reg.pulverizacoes:
            if p.tipo == 'pre_plantio':
                ws_total.cell(row=linha_atual, column=col_pulv, value=p.classe_produto)
                ws_total.cell(row=linha_atual, column=col_pulv+1, value=p.alvo)
                col_pulv += 2
        
        # Pós-emergência (1 a 7)
        for i in range(1, 8):
            for p in reg.pulverizacoes:
                if p.tipo == f'pos_{i}':
                    ws_total.cell(row=linha_atual, column=col_pulv, value=p.classe_produto)
                    ws_total.cell(row=linha_atual, column=col_pulv+1, value=p.alvo)
                    col_pulv += 2
                    break
        
        linha_atual += 1
    
    # ============================================================
    # AJUSTAR LARGURA DAS COLUNAS
    # ============================================================
    for ws in [ws_bd, ws_total]:
        for col in range(1, 100):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    # ============================================================
    # SALVAR E ENVIAR
    # ============================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        download_name=f'idr_soja_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/delete/<int:id>', methods=['POST'])
def delete_record(id):
    registro = FormularioSoja.query.get_or_404(id)
    db.session.delete(registro)
    db.session.commit()
    flash('Registro excluído com sucesso!', 'success')
    return redirect(url_for('view_records'))

@app.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit_record(id):
    registro = FormularioSoja.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # Atualizar campos (mesma lógica do POST do form)
            registro.numero_produtor = request.form.get('numero_produtor')
            registro.regiao = request.form.get('regiao')
            registro.municipio = request.form.get('municipio')
            registro.meso_idr = request.form.get('meso_idr')
            registro.area_soja = float(request.form.get('area_soja') or 0)
            registro.produtividade_media = float(request.form.get('produtividade_media') or 0)
            registro.cultivar = request.form.get('cultivar')
            registro.bt = request.form.get('bt')
            registro.data_plantio = request.form.get('data_plantio')
            registro.data_emergencia = request.form.get('data_emergencia')
            registro.houve_adversidade = request.form.get('houve_adversidade')
            registro.qual_adversidade = request.form.get('qual_adversidade')
            registro.nome_coletor = request.form.get('nome_coletor')
            registro.unidade_municipal = request.form.get('unidade_municipal')
            
            registro.conhecimento_mid = request.form.get('conhecimento_mid')
            registro.utiliza_mid = request.form.get('utiliza_mid')
            registro.conhecimento_mip = request.form.get('conhecimento_mip')
            registro.utiliza_mip = request.form.get('utiliza_mip')
            
            registro.herbicida_dessecacao_alvo = request.form.get('herbicida_dessecacao_alvo')
            registro.herbicida_dessecacao_aplicacoes = int(request.form.get('herbicida_dessecacao_aplicacoes') or 0)
            registro.herbicida_pre_alvo = request.form.get('herbicida_pre_alvo')
            registro.herbicida_pre_aplicacoes = int(request.form.get('herbicida_pre_aplicacoes') or 0)
            registro.herbicida_pos_alvo = request.form.get('herbicida_pos_alvo')
            registro.herbicida_pos_aplicacoes = int(request.form.get('herbicida_pos_aplicacoes') or 0)
            
            registro.tratamento_sementes = request.form.get('tratamento_sementes')
            registro.sal_mistura = request.form.get('sal_mistura')
            registro.controle_biologico = request.form.get('controle_biologico')
            
            registro.inoculacao_sementes = request.form.get('inoculacao_sementes')
            registro.forma_inoculacao = request.form.get('forma_inoculacao')
            registro.coinoculacao = request.form.get('coinoculacao')
            registro.co_mo = request.form.get('co_mo')
            registro.co_mo_aplicacao = request.form.get('co_mo_aplicacao')
            
            # Remover pulverizações antigas
            Pulverizacao.query.filter_by(formulario_id=registro.id).delete()

            # Pré-plantio com múltiplas classes
            if request.form.get('data_pre_plantio'):  # <--- PRECISA DE 12 ESPAÇOS NO INÍCIO!
                  classes_pre = request.form.getlist('classe_pre_plantio')
                  if classes_pre:
                      classe_pre_str = ', '.join(classes_pre)
                  else:
                      classe_pre_str = ''
                  
                  alvo_pre = request.form.get('alvo_pre_plantio')
                  
                  if classe_pre_str and alvo_pre:
                      pulv = Pulverizacao(
                          formulario_id=registro.id,
                          tipo='pre_plantio',
                          data=request.form.get('data_pre_plantio'),
                          classe_produto=classe_pre_str,
                          alvo=alvo_pre
                      )
                      db.session.add(pulv)
            
            for i in range(1, 8):
                data = request.form.get(f'data_pos_{i}')
                if data:
                    classes = request.form.getlist(f'classe_pos_{i}')
                    if classes:
                        classe_str = ', '.join(classes)
                    else:
                        classe_str = ''
                    
                    alvo = request.form.get(f'alvo_pos_{i}')
                    
                    if data and classe_str:
                        pulv = Pulverizacao(
                            formulario_id=registro.id,
                            tipo=f'pos_{i}',
                            data=data,
                            classe_produto=classe_str,
                            alvo=alvo
                        )
                        db.session.add(pulv)
            
            db.session.commit()
            flash('Registro atualizado com sucesso!', 'success')
            return redirect(url_for('view_record', id=registro.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erro ao atualizar: {str(e)}', 'danger')
    
    return render_template('edit_form.html', 
                      registro=registro, 
                      insetos=INSETOS_ALVO, 
                      doencas=DOENCAS_ALVO,
                      plantas=PLANTAS_DANINHAS,
                      acaros=ACAROS)
    
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)
