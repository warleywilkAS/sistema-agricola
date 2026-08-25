"""
export_excel.py
---------------
Gera o Excel de exportacao no formato EXATO do arquivo "Modelo_1.xlsx":

    Aba 1 "BD"        -> dados de referencia gerais (listas fixas:
                          doencas, pragas, cultivares, regionais/municipios,
                          plantas invasoras, classes de produto, etc.)
    Aba 2 "Total_Pr"  -> tabulacao com as 129 colunas, no mesmo layout
                          de cabecalho do modelo, preenchida com os
                          dados reais cadastrados no site.

Nao ha formatacao (cor, negrito, largura, congelar paineis etc.) - apenas
o conteudo, exatamente como solicitado.

Uso no app.py (sem mudancas):

    from export_excel import gerar_excel, orm_para_dict

    @app.route("/exportar_excel")
    def exportar_excel():
        todos = FormularioSoja.query.order_by(FormularioSoja.id).all()
        registros = [orm_para_dict(r) for r in todos]
        filepath = os.path.join("/tmp", "MesoIDR_Export.xlsx")
        gerar_excel(registros, filepath)
        return send_file(filepath, as_attachment=True,
                         download_name="MesoIDR_Exportacao.xlsx",
                         mimetype="application/vnd.openxmlformats-"
                                  "officedocument.spreadsheetml.sheet")

IMPORTANTE: o arquivo "bd_dados.json" (dados fixos da aba BD) precisa estar
na MESMA PASTA deste arquivo export_excel.py.
"""

from __future__ import annotations

import json
import os
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference

# ---------------------------------------------------------------------------
# Constantes de dominio (usadas apenas para reconhecer "alvo" texto -> praga/
# doenca ao montar o dicionario de cada registro)
# ---------------------------------------------------------------------------
REGIOES_IDR = [
    "Noroeste", "Norte", "Oeste", "Sudoeste",
    "Centro Sul", "Centro", "Metropolitana e Litoral",
]

PRAGAS = [
    "Lagarta da soja (Anticarsia gemmatalis)",
    "Lagarta das vagens (Spodoptera spp.)",
    "Lagarta falsa medideira (Chrysodeixis includens)",
    "Lagartas do grupo Heliothinae",
    "Percevejo barriga verde (Dichelops spp.)",
    "Percevejo marrom (Euschistus heros)",
    "Percevejo verde (Nezara viridula)",
    "Percevejo verde pequeno (Piezodorus guildinii)",
    "Broca dos ponteiros (Crocidosema aporema)",
    "Mosca Branca",
    "Outros insetos praga",
    "Tamandua da soja (Sternechus subsignatus)",
    "Tripes",
    "Vaquinhas (Diabrotica/ Cerotoma/ Colapsis)",
]

ACAROS = [
    "Acaro-rajado (Tetranychus urticae)",
    "Acaro-verde (Mononychellus planki)",
    "Acaro-branco (Polyphagotarsonemus latus)",
    "Acaros-vermelhos (Tetranychus spp.)",
    "Outros acaros",
]

DOENCAS_FUNGICAS = [
    "Antracnose (Colletotrichum truncatum)",
    "Cancro da haste (Diaporthe spp.)",
    "Ferrugem asiatica (Phakopsora pachyrhizi)",
    "Mancha alvo (Corynespora cassiicola)",
    "Mancha de cercospora (Cercospora kikuchii)",
    "Mancha olho-de-ra (Cercospora sojina)",
    "Mancha parda (Septoria glycines)",
    "Mela ou requeima (Rhizoctonia solani)",
    "Mofo branco (Sclerotinia sclerotiorum)",
    "Mildio (Peronospora manshurica)",
    "Oidio (Microsphaera diffusa)",
    "Outras Doencas Fungicas",
]
DOENCAS_BACT = [
    "Crestamento bacteriano (Pseudomonas savastanoi pv. glycinea)",
    "Fogo selvagem (Pseudomonas syringae pv. tabaci)",
    "Pustula bacteriana (Xanthomonas axonopodis pv. glycines)",
    "Mancha bacteriana marrom (Curtobacterium flaccumfaciens pv. flaccumfaciens)",
]
DOENCAS = DOENCAS_FUNGICAS + DOENCAS_BACT

N_PULV = 7

_LAGARTAS = [p for p in PRAGAS if "Lagarta" in p]
_PERCEVEJOS = [p for p in PRAGAS if "Percevejo" in p]
_PRAGAS_NOMEADAS = _LAGARTAS + _PERCEVEJOS
_OUTRAS_PRAGAS = [p for p in PRAGAS if p not in _PRAGAS_NOMEADAS]

_FERRUGEM = [d for d in DOENCAS_FUNGICAS if "Ferrugem" in d]
_MANCHA_ALVO = [d for d in DOENCAS_FUNGICAS if "Mancha alvo" in d]
_OIDIO = [d for d in DOENCAS_FUNGICAS if "Oidio" in d or "Oídio" in d]
_DOENCAS_NOMEADAS = _FERRUGEM + _MANCHA_ALVO + _OIDIO
_DEMAIS_FUNGICAS = [d for d in DOENCAS_FUNGICAS if d not in _DOENCAS_NOMEADAS]
_DOENCAS_MENOS_FERRUGEM = [d for d in DOENCAS if d not in _FERRUGEM]


def _norm(txt) -> str:
    """minusculas, sem acento, para comparacao tolerante de texto"""
    if not txt:
        return ""
    import unicodedata
    s = unicodedata.normalize("NFKD", str(txt))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.strip().lower()


def _match_lista(alvo, lista) -> bool:
    """True se o texto 'alvo' corresponde a algum item de 'lista' (comparacao
    tolerante: normaliza acentos/maiusculas e aceita substring)."""
    if not alvo:
        return False
    a = _norm(alvo)
    for item in lista:
        n = _norm(item)
        if a == n or n in a or a in n:
            return True
    return False


# ---------------------------------------------------------------------------
# Conversao ORM -> dict
# ---------------------------------------------------------------------------
def orm_para_dict(r) -> dict:
    """
    Converte FormularioSoja (com Pulverizacoes) para dict usado por gerar_excel().

    Modelo Pulverizacao:
      .tipo           -> 'dessecacao' | '1' | '2' ... '7'
      .data           -> string 'YYYY-MM-DD'
      .classe_produto -> ex. "Inseticida, Fungicida"
      .alvo           -> ex. "Lagarta da soja (Anticarsia gemmatalis), Ferrugem..."
    """
    pulvs: dict[str, Any] = {}
    for p in (r.pulverizacoes or []):
        pulvs[str(p.tipo).strip()] = p

    def _split(text: str) -> list[str]:
        if not text:
            return []
        return [x.strip() for x in text.replace("\n", ",").split(",") if x.strip()]

    def _dae(n: int):
        obj = pulvs.get(f"pos_{n}")
        if not obj or not obj.data or not r.data_emergencia:
            return None
        try:
            from datetime import datetime
            dp = datetime.strptime(obj.data[:10], "%Y-%m-%d").date()
            de = datetime.strptime(r.data_emergencia[:10], "%Y-%m-%d").date()
            return (dp - de).days
        except Exception:
            return None

    d: dict[str, Any] = {}

    # Identificacao
    d["N"]               = r.id
    d["Numero_Produtor"] = r.numero_produtor
    d["Meso_IDR"]        = r.meso_idr
    d["Regiao"]          = r.regiao
    d["Municipio"]       = r.municipio
    d["Area_Soja"]       = r.area_soja
    d["Cultivar"]        = r.cultivar
    d["Bt"]              = r.bt
    d["Produtividade"]   = r.produtividade_media
    d["Dt_Plantio"]      = r.data_plantio
    d["Adversidade"]     = r.qual_adversidade if r.houve_adversidade == "SIM" else None
    d["Sinistro"]        = r.houve_adversidade

    # Monitoramento
    d["Conhec_MID"]  = r.conhecimento_mid
    d["Utiliza_MID"] = r.utiliza_mid
    d["Conhec_MIP"]  = r.conhecimento_mip
    d["Utiliza_MIP"] = r.utiliza_mip

    # Plantas invasoras (ate 3 categorias hoje: dessecacao / pre / pos;
    # a 4a coluna do modelo fica em branco pois o site nao coleta esse dado)
    herbs = [
        ("Herbicida", r.herbicida_dessecacao_alvo, r.herbicida_dessecacao_aplicacoes),
        ("Herbicida", r.herbicida_pre_alvo,         r.herbicida_pre_aplicacoes),
        ("Herbicida", r.herbicida_pos_alvo,         r.herbicida_pos_aplicacoes),
        ("Herbicida", getattr(r, "herbicida_pos_ns_alvo", None), getattr(r, "herbicida_pos_ns_aplicacoes", None)),
    ]
    for i, (cl, alv, nap) in enumerate(herbs, start=1):
        d[f"Herb_Cl{i}"]  = cl if alv else None
        d[f"Herb_Alv{i}"] = alv
        d[f"Herb_Nap{i}"] = nap

    # Dessecacao (evento especifico, com data e ate 3 alvos)
    dess = pulvs.get("pre_plantio")
    d["Dess_Sim"] = "SIM" if dess else "NAO"
    d["Dess_Dt"]  = dess.data if dess else None
    d["Dess_Cl"]  = dess.classe_produto if dess else None
    dess_alvos = _split(dess.alvo) if dess else []
    for i in range(1, 4):
        d[f"Dess_Alv{i}"] = dess_alvos[i - 1] if i <= len(dess_alvos) else None

    # Pulverizacoes 1-7 (ate 5 classes/alvos por aplicacao)
    for n in range(1, N_PULV + 1):
        obj = pulvs.get(f"pos_{n}")
        alvos   = _split(obj.alvo)           if obj else []
        classes = _split(obj.classe_produto) if obj else []

        d[f"P{n}_DAE"]  = _dae(n)
        d[f"P{n}_Data"] = obj.data if obj else None

        for k in range(1, 6):
            d[f"P{n}_Cl{k}"]  = classes[k - 1] if k <= len(classes) else None
            d[f"P{n}_Alv{k}"] = alvos[k - 1]   if k <= len(alvos)   else None

    # Outras
    d["Tto_Semente"] = r.tratamento_sementes
    d["SAL_CB"]      = r.sal_mistura
    d["Ctrl_Biol"]   = r.controle_biologico

    # FBN / Inoculacao
    d["Inoc_Usa"]   = r.inoculacao_sementes
    d["Inoc_Forma"] = r.forma_inoculacao
    d["Coinoc"]     = r.coinoculacao
    d["CoMo_Usa"]   = r.co_mo
    d["CoMo_Forma"] = r.co_mo_aplicacao

    return d


# ---------------------------------------------------------------------------
# Aba "BD" - dados de referencia gerais (identico ao Modelo_1.xlsx)
# ---------------------------------------------------------------------------
def _carregar_bd_dados() -> list[list]:
    """Le o arquivo bd_dados.json (deve estar ao lado deste .py)."""
    caminho = os.path.join(os.path.dirname(os.path.abspath(__file__)), "bd_dados.json")
    with open(caminho, "r", encoding="utf-8") as f:
        return json.load(f)


def _build_BD(wb: Workbook):
    ws = wb.create_sheet("BD")
    linhas = _carregar_bd_dados()
    for ri, linha in enumerate(linhas, start=1):
        for ci, val in enumerate(linha, start=1):
            if val is not None:
                ws.cell(row=ri, column=ci, value=val)


# ---------------------------------------------------------------------------
# Layout da aba "Total_Pr" (129 colunas), identico ao Modelo_1.xlsx
# Cada item: (chave_no_dict, texto_do_cabecalho, formato)
# formato: "num" -> numero, "txt"/None -> texto
# ---------------------------------------------------------------------------
def _pulv_group_cols(n: int, texto_dae: str):
    cols = [
        (f"P{n}_DAE",  texto_dae, "num"),
        (f"P{n}_Data", "Data",    "txt"),
    ]
    for k in range(1, 6):
        cols += [
            (f"P{n}_Cl{k}",  "Classe do Produto", "txt"),
            (f"P{n}_Alv{k}", "Alvo",              "txt"),
        ]
    return cols


_TEXTO_DAE = {
    1: "1ª Pulverização   (DAE)",
    2: "2ª Pulverização (DAE)",
    3: "3ª Pulverização   (DAE)",
    4: "4ª Pulverização  (DAE)",
    5: "5ª Pulverização  (DAE)",
    6: "6ª Pulverização  (DAE)",
    7: "7ª Pulverização (DAE)",
}

_TITULO_TOTAL_PR = (
    "PLANILHA TABULAÇÃO DADOS QUESTIONÁRIOS APLICAÇÃO DEFENSIVOS PARA "
    "CONTROLE PRAGAS E DOENÇAS_PR_SAFRA 19_20_V1"
)

# Colunas 1-4: sem "chave" de grupo (label fica direto na linha 3, sem
# subcabecalho na linha 4), exatamente como no modelo.
_COLS_INICIAIS = [
    (None,               None,     None),   # col 1 - sem cabecalho (numero sequencial)
    ("_TABELA",          "Tabela", "txt"),   # col 2
    ("Numero_Produtor",  "N° P",   "txt"),   # col 3
    (None,               "Ordem ", None),    # col 4 - sem subcabecalho (numero sequencial)
]

_ID_COLS = [
    ("Meso_IDR",      "Meso_IDR",                    "txt"),
    ("Regiao",        "Região",                      "txt"),
    ("Municipio",     "Município",                   "txt"),
    ("Area_Soja",     "Área com  Soja (ha)",          "num"),
    ("Cultivar",      "Cultivar",                     "txt"),
    ("Bt",            "Bt",                           "txt"),
    ("Produtividade", "Produtividade Média (sc/ha)",  "num"),
    ("Dt_Plantio",    "Data Plantio",                 "txt"),
    ("Adversidade",   "Adversidade",                  "txt"),
    ("Sinistro",      "Sinistro",                     "txt"),
]

_MID_COLS = [
    ("Conhec_MID",  "Conhec. MID",  "txt"),
    ("Utiliza_MID", "Utiliza MID",  "txt"),
    ("Conhec_MIP",  "Conhec. MIP",  "txt"),
    ("Utiliza_MIP", "Utiliza MIP",  "txt"),
]

# 4 blocos de (Classe do Produto / Alvo / N° Aplicações)
_HERB_COLS = []
for _i in range(1, 5):
    _HERB_COLS += [
        (f"Herb_Cl{_i}",  "Classe do Produto", "txt"),
        (f"Herb_Alv{_i}", "Alvo",              "txt"),
        (f"Herb_Nap{_i}", "N° Aplicações",     "num"),
    ]

_DESS_COLS = [
    ("Dess_Sim",  " Pulverização na Dessecação  ", "txt"),
    ("Dess_Dt",   "Data",                          "txt"),
    ("Dess_Cl",   "Classe do Produto",             "txt"),
    ("Dess_Alv1", "Alvo_1",                        "txt"),
    ("Dess_Alv2", "Alvo_2",                        "txt"),
    ("Dess_Alv3", "Alvo_3",                        "txt"),
]

_OUTRAS_COLS = [
    ("Tto_Semente", "Tratamento de Semente",           "txt"),
    ("SAL_CB",      "Utilização de SAL + Inseticida",  "txt"),
    ("Ctrl_Biol",   "Utilizou Controle Biológico",     "txt"),
]

_INOC_COLS = [
    ("Inoc_Usa",   "Utiliza Inoculação", "txt"),
    ("Inoc_Forma", "Forma Inoculação",   "txt"),
    ("Coinoc",     "Coinoculação",       "txt"),
    ("CoMo_Usa",   "Utiliza Co e Mo",    "txt"),
    ("CoMo_Forma", "Forma Co e Mo",      "txt"),
]

# grupos = (texto_do_grupo_ou_None, lista_de_colunas)
GRUPOS_TOTAL_PR = [
    (None,                                                     _COLS_INICIAIS),
    (None,                                                     _ID_COLS),
    ("CONHECIMENTO MONITORAMENTO",                             _MID_COLS),
    ("3_Informação Plantas Invasoras",                         _HERB_COLS),
    ("4.0_INFORMAÇÃO _PULVERIZAÇÃO DESSECAÇÃO",                _DESS_COLS),
    ("4.1_INFORMAÇÃO _PRIMEIRA PULVERIZAÇÃO APÓS EMERGÊNCIA",  _pulv_group_cols(1, _TEXTO_DAE[1])),
    ("4.2_INFORMAÇÃO _SEGUNDA PULVERIZAÇÃO APÓS EMERGÊNCIA",   _pulv_group_cols(2, _TEXTO_DAE[2])),
    ("4.3_INFORMAÇÃO _TERCEIRA PULVERIZAÇÃO APÓS EMERGÊNCIA",  _pulv_group_cols(3, _TEXTO_DAE[3])),
    ("4.4_INFORMAÇÃO _QUARTA PULVERIZAÇÃO APÓS EMERGÊNCIA",    _pulv_group_cols(4, _TEXTO_DAE[4])),
    ("4.5_INFORMAÇÃO _QUINTA PULVERIZAÇÃO APÓS EMERGÊNCIA",    _pulv_group_cols(5, _TEXTO_DAE[5])),
    ("4.6_INFORMAÇÃO _SEXTA PULVERIZAÇÃO APÓS EMERGÊNCIA",     _pulv_group_cols(6, _TEXTO_DAE[6])),
    ("4.7_INFORMAÇÃO _SÉTIMA PULVERIZAÇÃO APÓS EMERGÊNCIA",    _pulv_group_cols(7, _TEXTO_DAE[7])),
    ("5.OUTRAS INFORMAÇÕES",                                   _OUTRAS_COLS),
    ("6.INOCULAÇÃO",                                           _INOC_COLS),
]

ALL_COLS: list[tuple] = []
for _, cols in GRUPOS_TOTAL_PR:
    ALL_COLS.extend(cols)

# coluna 129 (DY) fica em branco, igual ao modelo
ALL_COLS.append((None, None, None))

_CI: dict[str, int] = {}
for _idx, (key, _label, _fmt) in enumerate(ALL_COLS, start=1):
    if key:
        _CI[key] = _idx


def _build_total_pr(wb: Workbook, registros: list[dict]):
    ws = wb.create_sheet("Total_Pr")
    nc = len(ALL_COLS)

    # Linha 1: titulo (somente na coluna D, igual ao modelo)
    ws.cell(row=1, column=4, value=_TITULO_TOTAL_PR)

    # Linha 3: cabecalhos de grupo (so na primeira coluna de cada grupo)
    # Linha 4: subcabecalhos (uma celula por coluna)
    col = 1
    for grupo_texto, cols in GRUPOS_TOTAL_PR:
        if grupo_texto is not None:
            ws.cell(row=3, column=col, value=grupo_texto)
        for key, label, _fmt in cols:
            if label is not None:
                # colunas iniciais (Tabela/N°P) tem o texto na propria linha 3
                if grupo_texto is None and cols is _COLS_INICIAIS:
                    ws.cell(row=3, column=col, value=label)
                else:
                    ws.cell(row=4, column=col, value=label)
            col += 1

    # Linha 5 fica em branco (igual ao modelo); dados comecam na linha 6
    linha_inicial = 6
    for i, reg in enumerate(registros):
        ri = linha_inicial + i
        ws.cell(row=ri, column=1, value=i + 1)      # col 1 - sequencial
        ws.cell(row=ri, column=2, value="TB1.")      # col 2 - Tabela
        ws.cell(row=ri, column=4, value=i + 1)       # col 4 - Ordem
        for key, _label, _fmt in ALL_COLS:
            if not key or key in ("_TABELA",):
                continue
            ci = _CI.get(key)
            if ci:
                ws.cell(row=ri, column=ci, value=reg.get(key))


# ---------------------------------------------------------------------------
# Aba "Médias_Geral" - indicadores agregados (calculados em Python direto a
# partir dos registros, no lugar de reproduzir as formulas originais que
# dependiam da aba auxiliar "Contagem_Pragas" de 534 colunas).
#
# LIMITACOES CONHECIDAS (o site hoje nao coleta esses dados, entao ficam
# zerados/ausentes ate que o formulario seja ampliado):
#   - Quebra "Folha Larga / Folha Estreita" dos herbicidas (o modelo original
#     usa uma classificacao escolhida manualmente no questionario, que nao
#     existe no banco atual).
#   - 4a categoria de "Plantas Invasoras" (o site so registra 3: dessecacao,
#     pre-emergente e pos-emergente).
# ---------------------------------------------------------------------------
REGIOES_COLUNAS = [None] + REGIOES_IDR  # None = "PARANA" (todo o estado)


def _filtra(registros: list[dict], regiao: str | None = None, bt: str | None = None):
    out = registros
    if regiao is not None:
        # "regiao" aqui refere-se a MESORREGIAO (Noroeste, Norte, Oeste...),
        # que fica salva no campo Meso_IDR de cada registro. O campo "Regiao"
        # do registro guarda a Unidade Regional (URE, ex: Apucarana, Cascavel),
        # que é uma divisão diferente (mais fina) e não deve ser usada aqui.
        out = [r for r in out if _norm(r.get("Meso_IDR")) == _norm(regiao)]
    if bt is not None:
        out = [r for r in out if _norm(r.get("Bt")) == _norm(bt)]
    return out


def _aplicacoes_alvo(reg: dict, lista_alvo: list[str]):
    """Lista de DAE (dias apos emergencia) de cada pulverizacao (1-7) que
    tenha atingido pelo menos um alvo de 'lista_alvo'."""
    daes = []
    for n in range(1, N_PULV + 1):
        alvos = [reg.get(f"P{n}_Alv{k}") for k in range(1, 6)]
        if any(_match_lista(a, lista_alvo) for a in alvos if a):
            dae = reg.get(f"P{n}_DAE")
            if dae is not None:
                daes.append(dae)
    return daes


def _media(lst):
    lst = [x for x in lst if x is not None]
    return (sum(lst) / len(lst)) if lst else None


def _bloco_alvo(registros: list[dict], lista_alvo: list[str]) -> dict:
    """Calcula os 10 indicadores padrao (COM/SEM aplicacao, %, n aplicacoes,
    DAE medio/primeira/menor/maior) para um alvo (praga, doenca, etc.)."""
    total = len(registros)
    com = sem = 0
    n_aplic_all, n_aplic_aplic = [], []
    dae_pool, dae_primeira = [], []
    for reg in registros:
        daes = _aplicacoes_alvo(reg, lista_alvo)
        n_aplic_all.append(len(daes))
        if daes:
            com += 1
            n_aplic_aplic.append(len(daes))
            dae_pool.extend(daes)
            dae_primeira.append(min(daes))
        else:
            sem += 1
    return {
        "com": com,
        "sem": sem,
        "pct_com": (com / total) if total else None,
        "pct_sem": (sem / total) if total else None,
        "n_aplic_total": _media(n_aplic_all),
        "n_aplic_aplicantes": _media(n_aplic_aplic),
        "dae_medio": _media(dae_pool),
        "dae_primeira": _media(dae_primeira),
        "menor_dae_primeira": min(dae_primeira) if dae_primeira else None,
        "maior_dae_primeira": max(dae_primeira) if dae_primeira else None,
    }


_METRICAS_ALVO = [
    ("N° Questionários COM Aplicação {N}", "com"),
    ("N° Questionários SEM Aplicação {N}", "sem"),
    ("% Questionários COM Aplicação {N}", "pct_com"),
    ("% Questionários SEM Aplicação {N}", "pct_sem"),
    ("N° aplicações para {N} Total", "n_aplic_total"),
    ("N° aplicações para {N} Aplicantes", "n_aplic_aplicantes"),
    ("DAE_Médio aplicação {N}", "dae_medio"),
    ("DAE_Primeira aplicação {N}", "dae_primeira"),
    ("Menor DAE_Primeira Aplicação {N}", "menor_dae_primeira"),
    ("Maior DAE_Primeira Aplicação {N}", "maior_dae_primeira"),
]

# blocos de alvo, na mesma ordem do arquivo original
_BLOCOS_ALVO_PRAGAS = [
    ("PRAGAS", PRAGAS),
    ("LAGARTAS", _LAGARTAS),
    ("ANTICARSIA GEMMATALIS", [p for p in PRAGAS if "Anticarsia" in p]),
    ("SPODOPTERA ssp.", [p for p in PRAGAS if "Spodoptera" in p]),
    ("CHRYSODEIXIS INCLUDENS", [p for p in PRAGAS if "Chrysodeixis" in p]),
    ("Grupo HELIOTHINAE", [p for p in PRAGAS if "Heliothinae" in p]),
    ("PERCEVEJOS", _PERCEVEJOS),
    ("DICHELOPS sp.", [p for p in PRAGAS if "Dichelops" in p]),
    ("EUSCHISTUS HEROS", [p for p in PRAGAS if "Euschistus" in p]),
    ("NEZARA VIRIDULA", [p for p in PRAGAS if "Nezara" in p]),
    ("PIEZODORUS GUILDINI", [p for p in PRAGAS if "Piezodorus" in p]),
    ("OUTRAS PRAGAS", _OUTRAS_PRAGAS),
    ("ÁCAROS", ACAROS),
    ("OUTRAS PRAGAS + Ácaros", _OUTRAS_PRAGAS + ACAROS),
]

_BLOCOS_ALVO_DOENCAS = [
    ("DOENÇAS", DOENCAS),
    ("FERRUGEM", _FERRUGEM),
    ("MANCHA ALVO", _MANCHA_ALVO),
    ("OÍDIO", _OIDIO),
    ("DEMAIS DOENÇAS FÚNGICAS", _DEMAIS_FUNGICAS),
    ("DOENÇAS BACTERIANAS", DOENCAS_BACT),
    ("TODAS DOENÇAS MENOS FERRUGEM", _DOENCAS_MENOS_FERRUGEM),
]

_BLOCOS_ALVO = _BLOCOS_ALVO_PRAGAS + _BLOCOS_ALVO_DOENCAS


def _bloco_simples(registros, campo, valor_sim="SIM", valor_nao="NAO"):
    """Contagem COM/SEM/SEM_RESPOSTA para um campo de resposta SIM/NAO."""
    com = sem = sem_resposta = 0
    for reg in registros:
        v = _norm(reg.get(campo))
        if v == _norm(valor_sim):
            com += 1
        elif v == _norm(valor_nao):
            sem += 1
        else:
            sem_resposta += 1
    total = len(registros)
    return {
        "respondido": com + sem,
        "com": com,
        "sem": sem,
        "sem_resposta": sem_resposta,
        "pct_com": (com / total) if total else None,
        "pct_sem": (sem / total) if total else None,
    }


def _calcular_medias_geral(registros: list[dict]):
    """
    Calcula todas as linhas da Medias_Geral, independente do Excel.
    Retorna (col_map, linhas):
      col_map = [(regiao_ou_None, bt_ou_None), ...]  # 24 combinacoes, na ordem das colunas
      linhas  = [(label, [valores alinhados a col_map]) , (None, None) para linha em branco, ...]
    """
    col_map = []  # (regiao_ou_None, bt_ou_None) -- sem indice de coluna, calculado depois so pro Excel
    for regiao in REGIOES_COLUNAS:
        col_map.append((regiao, None))
        col_map.append((regiao, "SIM"))
        col_map.append((regiao, "NAO"))

    linhas = []  # (label, [valores]) ou (None, None) = linha em branco

    def escreve_linha(label, valores_por_regiao_bt):
        """valores_por_regiao_bt: dict {(regiao,bt): valor}"""
        linhas.append((label, [valores_por_regiao_bt.get(chave) for chave in col_map]))

    def pula_linha(n=1):
        for _ in range(n):
            linhas.append((None, None))

    def para_cada_grupo(fn_calc):
        """fn_calc(regs_filtrados) -> dict de metricas; roda para todas as
        combinacoes de regiao/bt e devolve {(regiao,bt): dict}"""
        out = {}
        for regiao, bt in col_map:
            regs = _filtra(registros, regiao=regiao, bt=bt)
            out[(regiao, bt)] = fn_calc(regs)
        return out

    # ---------------- Bloco A: identificacao geral ----------------
    resultados_id = para_cada_grupo(lambda regs: {
        "n_aplicados": len(regs),
        "area_soja": sum(r.get("Area_Soja") or 0 for r in regs),
        "area_media": _media([r.get("Area_Soja") for r in regs]),
        "produtividade": _media([r.get("Produtividade") for r in regs]),
        "com_sinistro": sum(1 for r in regs if _norm(r.get("Sinistro")) == "sim"),
        "sem_sinistro": sum(1 for r in regs if _norm(r.get("Sinistro")) == "nao"),
        "area_com_sinistro": sum((r.get("Area_Soja") or 0) for r in regs if _norm(r.get("Sinistro")) == "sim"),
    })

    def _get(chave):
        return {k: v[chave] for k, v in resultados_id.items()}

    escreve_linha("N° Questionários Aplicados", _get("n_aplicados"))
    n_total_geral = resultados_id[(None, None)]["n_aplicados"]

    def _pct(vals_key, base=n_total_geral):
        out = {}
        for k, v in resultados_id.items():
            n = v["n_aplicados"]
            out[k] = (n / n_total_geral) if n_total_geral else None
        return out

    escreve_linha("Percentual Questionários", _pct("n_aplicados"))
    escreve_linha("N° Questionários COM Relato de SINISTRO", _get("com_sinistro"))
    escreve_linha("% Questionários COM Relato de SINISTRO",
                  {k: (v["com_sinistro"] / v["n_aplicados"]) if v["n_aplicados"] else None
                   for k, v in resultados_id.items()})
    escreve_linha("N° Questionários SEM Relato de SINISTRO", _get("sem_sinistro"))
    escreve_linha("% Questionários SEM Relato de SINISTRO",
                  {k: (v["sem_sinistro"] / v["n_aplicados"]) if v["n_aplicados"] else None
                   for k, v in resultados_id.items()})
    escreve_linha("Área_SOJA (ha)", _get("area_soja"))
    escreve_linha("Área média cultivada (ha)", _get("area_media"))
    escreve_linha("Área_Soja COM Relato SINISTRO", _get("area_com_sinistro"))
    escreve_linha("% área_Soja COM Relato SINISTRO",
                  {k: (v["area_com_sinistro"] / v["area_soja"]) if v["area_soja"] else None
                   for k, v in resultados_id.items()})
    escreve_linha("Produtividade (sc/ha)", _get("produtividade"))
    pula_linha()

    # ---------------- Blocos B/C/D: pragas, acaros, doencas ----------------
    for nome, lista in _BLOCOS_ALVO:
        resultados = para_cada_grupo(lambda regs, lst=lista: _bloco_alvo(regs, lst))
        for template, chave in _METRICAS_ALVO:
            escreve_linha(template.format(N=nome), {k: v[chave] for k, v in resultados.items()})
        pula_linha()

    # ---------------- Bloco E: totais gerais de aplicacao ----------------
    res_pragas = para_cada_grupo(lambda regs: _bloco_alvo(regs, PRAGAS))
    res_doencas = para_cada_grupo(lambda regs: _bloco_alvo(regs, DOENCAS))
    escreve_linha("Número Total Aplicação Inseticida",
                  {k: v["n_aplic_total"] for k, v in res_pragas.items()})
    escreve_linha("Número Total Aplicação Total Fungicida",
                  {k: v["n_aplic_total"] for k, v in res_doencas.items()})
    escreve_linha("Número Total Aplicação Total (Inseticida + Fungicida)",
                  {k: (res_pragas[k]["n_aplic_total"] or 0) + (res_doencas[k]["n_aplic_total"] or 0)
                   for k in res_pragas})
    pula_linha()

    # ---------------- Bloco F: tratamento semente / SAL / ctrl biologico ----------------
    def _bloco_campo(campo, valor_sim="SIM", valor_nao="NAO"):
        return para_cada_grupo(lambda regs: _bloco_simples(regs, campo, valor_sim, valor_nao))

    for campo, titulo in [
        ("Tto_Semente", "TRATAMENTO SEMENTE"),
        ("SAL_CB", "Utilização de SAL para controle percevejos"),
        ("Ctrl_Biol", "CONTROLE BIOLÓGICO"),
    ]:
        r = _bloco_campo(campo)
        escreve_linha(f"N° Questionários {titulo}", {k: v["respondido"] for k, v in r.items()})
        escreve_linha(f"N° Questionários COM Utilização de {titulo}", {k: v["com"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM Utilização {titulo}", {k: v["sem"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM RESPOSTA {titulo}", {k: v["sem_resposta"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários COM {titulo}", {k: v["pct_com"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários SEM {titulo}", {k: v["pct_sem"] for k, v in r.items()})
        pula_linha()

    # Inseticida na dessecação (classe do produto contem "Inseticida")
    r_ins_dess = para_cada_grupo(lambda regs: _bloco_simples_texto(regs, "Dess_Cl", "Inseticida"))
    escreve_linha("N° Questionários COM utilização de INSETICIDA NA DESSECAÇÃO",
                  {k: v["com"] for k, v in r_ins_dess.items()})
    escreve_linha("N° Questionários SEM utilização de INSETICIDA NA DESSECAÇÃO",
                  {k: v["sem"] for k, v in r_ins_dess.items()})
    escreve_linha("Percentual Questionários COM utilização de INSETICIDA NA DESSECAÇÃO",
                  {k: v["pct_com"] for k, v in r_ins_dess.items()})
    escreve_linha("Percentual de Questionários SEM utilização de INSETICIDA NA DESSECAÇÃO",
                  {k: v["pct_sem"] for k, v in r_ins_dess.items()})
    pula_linha()

    # ---------------- Bloco G: conhecimento/uso MID e MIP ----------------
    for campo, titulo in [
        ("Conhec_MID", "CONHECIMENTO MID"),
        ("Utiliza_MID", "USO MID"),
        ("Conhec_MIP", "CONHECIMENTO MIP"),
        ("Utiliza_MIP", "USO MIP"),
    ]:
        r = _bloco_campo(campo)
        escreve_linha(f"N° Questionários {titulo}", {k: v["respondido"] for k, v in r.items()})
        escreve_linha(f"N° Questionários COM {titulo}", {k: v["com"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM {titulo}", {k: v["sem"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM RESPOSTA {titulo}", {k: v["sem_resposta"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários COM {titulo}", {k: v["pct_com"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários SEM {titulo}", {k: v["pct_sem"] for k, v in r.items()})
        pula_linha()

    # ---------------- Bloco H: herbicidas (COM/SEM aplicacao, geral) ----------------
    # OBS: a quebra "Folha Larga/Folha Estreita" do arquivo original depende de
    # uma classificacao manual que o site ainda nao coleta - por isso essas
    # sub-linhas nao aparecem aqui (ver observacao no topo do arquivo).
    for campo_alvo, campo_nap, titulo in [
        ("Herb_Alv1", "Herb_Nap1", "Herbicida não seletivo na dessecação"),
        ("Herb_Alv2", "Herb_Nap2", "Herbicida Pré emergente"),
        ("Herb_Alv3", "Herb_Nap3", "Herbicida Pós emergente"),
    ]:
        r = para_cada_grupo(lambda regs, c=campo_alvo: _bloco_simples_preenchido(regs, c))
        escreve_linha(f"N° Questionários COM aplicação {titulo}", {k: v["com"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM aplicação {titulo}", {k: v["sem"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários COM Aplicação {titulo}", {k: v["pct_com"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários SEM Aplicação {titulo}", {k: v["pct_sem"] for k, v in r.items()})
        pula_linha()

    # ---------------- Bloco I: inoculacao / coinoculacao / Co e Mo ----------------
    for campo, titulo in [
        ("Inoc_Usa", "USO INOCULAÇÃO"),
        ("Coinoc", "COINOCULAÇÃO"),
        ("CoMo_Usa", "USO Co Mo"),
    ]:
        r = _bloco_campo(campo)
        escreve_linha(f"N° Questionários RESPOSTA {titulo}", {k: v["respondido"] for k, v in r.items()})
        escreve_linha(f"N° Questionários COM {titulo}", {k: v["com"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM {titulo}", {k: v["sem"] for k, v in r.items()})
        escreve_linha(f"N° Questionários SEM RESPOSTA {titulo}", {k: v["sem_resposta"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários COM {titulo}", {k: v["pct_com"] for k, v in r.items()})
        escreve_linha(f"Percentual Questionários SEM {titulo}", {k: v["pct_sem"] for k, v in r.items()})
        pula_linha()

    # forma de inoculacao (categorias de texto livre)
    r_forma = para_cada_grupo(lambda regs: {
        cat: sum(1 for r in regs if cat_lower in _norm(r.get("Inoc_Forma")))
        for cat, cat_lower in [
            ("industrial", "industrial"),
            ("caixa", "caixa"),
            ("misturador", "misturador"),
            ("betoneira", "betoneira"),
            ("lona", "lona"),
            ("sulco", "sulco"),
        ]
    })
    for cat, titulo in [
        ("industrial", "INOCULAÇÃO INDUSTRIAL"),
        ("caixa", "INOCULAÇÃO CAIXA PLANTADEIRA"),
        ("misturador", "INOCULAÇÃO MISTURADOR SEMENTE"),
        ("betoneira", "INOCULAÇÃO BETONEIRA"),
        ("lona", "INOCULAÇÃO LONA"),
        ("sulco", "INOCULAÇÃO SULCO"),
    ]:
        escreve_linha(f"N° Questionários {titulo}", {k: v[cat] for k, v in r_forma.items()})
        escreve_linha(f"Percentual Questionários COM USO {titulo}",
                      {k: (v[cat] / n_total_geral) if n_total_geral else None for k, v in r_forma.items()})

    # Co e Mo: forma (semente/foliar)
    r_comomo = para_cada_grupo(lambda regs: {
        "semente": sum(1 for r in regs if "semente" in _norm(r.get("CoMo_Forma"))),
        "foliar": sum(1 for r in regs if "foliar" in _norm(r.get("CoMo_Forma"))),
    })
    pula_linha()
    escreve_linha("N° Questionários Co Mo SEMENTE", {k: v["semente"] for k, v in r_comomo.items()})
    escreve_linha("Percentual Questionários COM USO Co Mo SEMENTE",
                  {k: (v["semente"] / n_total_geral) if n_total_geral else None for k, v in r_comomo.items()})
    escreve_linha("N° Questionários Co Mo FOLIAR", {k: v["foliar"] for k, v in r_comomo.items()})
    escreve_linha("Percentual Questionários COM USO Co Mo FOLIAR",
                  {k: (v["foliar"] / n_total_geral) if n_total_geral else None for k, v in r_comomo.items()})

    return col_map, linhas


_BUCKETS_APLIC = [0, 1, 2, 3, 4, 5, 6, 7]

_BUCKETS_DAE = ["Não Aplicou", "Até 25", "26 a 40", "41 a 60", "61 a 80", "Mais de 80"]


def _bucket_dae(dae):
    if dae is None:
        return "Não Aplicou"
    if dae <= 25:
        return "Até 25"
    if dae <= 40:
        return "26 a 40"
    if dae <= 60:
        return "41 a 60"
    if dae <= 80:
        return "61 a 80"
    return "Mais de 80"


def _calcular_especies(registros: list[dict], especies: list[tuple]):
    """especies: lista de (nome, lista_alvo). Retorna col_map + {nome: [nº que aplicou, por col_map]}"""
    col_map = [(regiao, bt) for regiao in REGIOES_COLUNAS for bt in (None, "SIM", "NAO")]
    resultado = {nome: [] for nome, _lista in especies}
    for regiao, bt in col_map:
        regs = _filtra(registros, regiao=regiao, bt=bt)
        for nome, lista in especies:
            resultado[nome].append(_bloco_alvo(regs, lista)["com"])
    return col_map, resultado


def _calcular_grafico_mip(registros: list[dict]):
    """Dados para os graficos da aba Tabelas_Graficos_MIP: numero medio de
    aplicacoes por grupo de praga, composicao de especies de lagarta e de
    percevejo, e os histogramas (aplicacoes/DAE) de Lagartas e Percevejos."""
    col_map = [(regiao, bt) for regiao in REGIOES_COLUNAS for bt in (None, "SIM", "NAO")]

    grupos = [
        ("Lagartas", _LAGARTAS),
        ("Percevejos", _PERCEVEJOS),
        ("Outras Pragas + Ácaros", _OUTRAS_PRAGAS + ACAROS),
    ]
    media_aplicantes = []
    for _nome, lista in grupos:
        res = _calcular_distribuicao_alvo(registros, lista)
        media_aplicantes.append(res["aplicacoes"]["media_aplicantes"])

    especies_lagarta = [
        ("Anticarsia gemmatalis", [p for p in PRAGAS if "Anticarsia" in p]),
        ("Spodoptera spp.", [p for p in PRAGAS if "Spodoptera" in p]),
        ("Chrysodeixis includens", [p for p in PRAGAS if "Chrysodeixis" in p]),
        ("Grupo Heliothinae", [p for p in PRAGAS if "Heliothinae" in p]),
    ]
    _, r_esp_lag = _calcular_especies(registros, especies_lagarta)

    especies_percevejo = [
        ("Dichelops sp.", [p for p in PRAGAS if "Dichelops" in p]),
        ("Euschistus heros", [p for p in PRAGAS if "Euschistus" in p]),
        ("Nezara viridula", [p for p in PRAGAS if "Nezara" in p]),
        ("Piezodorus guildinii", [p for p in PRAGAS if "Piezodorus" in p]),
    ]
    _, r_esp_perc = _calcular_especies(registros, especies_percevejo)

    hist_lagartas = _calcular_distribuicao_alvo(registros, _LAGARTAS)
    hist_percevejos = _calcular_distribuicao_alvo(registros, _PERCEVEJOS)

    return {
        "col_map": col_map,
        "grupos_labels": [nome for nome, _lista in grupos],
        "media_aplicantes": media_aplicantes,
        "especies_lagarta": r_esp_lag,
        "especies_percevejo": r_esp_perc,
        "hist_lagartas_aplicacoes": hist_lagartas["aplicacoes"],
        "hist_lagartas_dae": hist_lagartas["dae"],
        "hist_percevejos_aplicacoes": hist_percevejos["aplicacoes"],
        "hist_percevejos_dae": hist_percevejos["dae"],
    }


def _calcular_grafico_mid(registros: list[dict]):
    """Dados para os graficos da aba Tabelas_Graficos_MID: numero medio de
    aplicacoes por grupo de doenca, composicao percentual das doencas, e o
    histograma (aplicacoes/DAE) das Doenças em geral."""
    col_map = [(regiao, bt) for regiao in REGIOES_COLUNAS for bt in (None, "SIM", "NAO")]

    grupos = [
        ("Ferrugem", _FERRUGEM),
        ("Mancha Alvo", _MANCHA_ALVO),
        ("Oídio", _OIDIO),
        ("Demais Fúngicas", _DEMAIS_FUNGICAS),
        ("Bacterianas", DOENCAS_BACT),
    ]
    media_aplicantes = []
    for _nome, lista in grupos:
        res = _calcular_distribuicao_alvo(registros, lista)
        media_aplicantes.append(res["aplicacoes"]["media_aplicantes"])

    _, r_composicao = _calcular_especies(registros, grupos)

    hist_doencas = _calcular_distribuicao_alvo(registros, DOENCAS)

    return {
        "col_map": col_map,
        "grupos_labels": [nome for nome, _lista in grupos],
        "media_aplicantes": media_aplicantes,
        "composicao_doencas": r_composicao,
        "hist_doencas_aplicacoes": hist_doencas["aplicacoes"],
        "hist_doencas_dae": hist_doencas["dae"],
    }


def _calcular_distribuicao_alvo(registros: list[dict], lista_alvo: list[str]):
    """
    Reproduz a logica da aba 'Lagartas'/'Percevejos'/etc: para um alvo (lista
    de pragas/doencas), calcula 2 histogramas por regiao x Bt/Nao Bt/Total:
      - frequencia do NUMERO de aplicacoes (0 a 7)
      - frequencia da faixa de DAE da 1a aplicacao (Nao aplicou / Ate 25 / ...)
    """
    col_map = [(regiao, bt) for regiao in REGIOES_COLUNAS for bt in (None, "SIM", "NAO")]

    aplic_contagem = {b: [] for b in _BUCKETS_APLIC}
    aplic_percentual = {b: [] for b in _BUCKETS_APLIC}
    aplic_media_aplicantes = []
    aplic_media_total = []

    dae_contagem = {b: [] for b in _BUCKETS_DAE}
    dae_percentual = {b: [] for b in _BUCKETS_DAE}
    dae_media_aplicantes = []
    dae_media_total = []

    for regiao, bt in col_map:
        regs = _filtra(registros, regiao=regiao, bt=bt)
        total = len(regs)

        n_aplic_list = []
        dae_primeira_list = []
        for reg in regs:
            daes = _aplicacoes_alvo(reg, lista_alvo)
            n_aplic_list.append(min(len(daes), 7))
            dae_primeira_list.append(min(daes) if daes else None)

        for b in _BUCKETS_APLIC:
            n = sum(1 for x in n_aplic_list if x == b)
            aplic_contagem[b].append(n)
            aplic_percentual[b].append((n / total) if total else None)

        aplicantes = [x for x in n_aplic_list if x > 0]
        aplic_media_aplicantes.append(_media(aplicantes))
        aplic_media_total.append(_media(n_aplic_list))

        for b in _BUCKETS_DAE:
            n = sum(1 for d in dae_primeira_list if _bucket_dae(d) == b)
            dae_contagem[b].append(n)
            dae_percentual[b].append((n / total) if total else None)

        daes_aplicantes = [d for d in dae_primeira_list if d is not None]
        dae_media_aplicantes.append(_media(daes_aplicantes))
        dae_media_total.append(_media(dae_primeira_list))

    return {
        "col_map": col_map,
        "aplicacoes": {
            "buckets": _BUCKETS_APLIC,
            "contagem": aplic_contagem,
            "percentual": aplic_percentual,
            "media_aplicantes": aplic_media_aplicantes,
            "media_total": aplic_media_total,
        },
        "dae": {
            "buckets": _BUCKETS_DAE,
            "contagem": dae_contagem,
            "percentual": dae_percentual,
            "media_aplicantes": dae_media_aplicantes,
            "media_total": dae_media_total,
        },
    }


def _calcular_percentual_campo(registros: list[dict], campo: str, valor_sim="SIM", valor_nao="NAO"):
    """Roda _bloco_simples() para cada combinacao de regiao x Bt/Nao Bt/Total."""
    col_map = [(regiao, bt) for regiao in REGIOES_COLUNAS for bt in (None, "SIM", "NAO")]
    chaves = ["respondido", "com", "sem", "sem_resposta", "pct_com", "pct_sem"]
    resultado = {k: [] for k in chaves}
    for regiao, bt in col_map:
        regs = _filtra(registros, regiao=regiao, bt=bt)
        r = _bloco_simples(regs, campo, valor_sim, valor_nao)
        for k in chaves:
            resultado[k].append(r[k])
    return col_map, resultado


def _calcular_percentual_texto(registros: list[dict], campo: str, contem: str):
    """Roda _bloco_simples_texto() para cada combinacao de regiao x Bt/Nao Bt/Total."""
    col_map = [(regiao, bt) for regiao in REGIOES_COLUNAS for bt in (None, "SIM", "NAO")]
    chaves = ["com", "sem", "pct_com", "pct_sem"]
    resultado = {k: [] for k in chaves}
    for regiao, bt in col_map:
        regs = _filtra(registros, regiao=regiao, bt=bt)
        r = _bloco_simples_texto(regs, campo, contem)
        for k in chaves:
            resultado[k].append(r[k])
    return col_map, resultado


def _calcular_categoria_texto(registros: list[dict], campo: str, categorias: list[tuple]):
    """categorias: lista de (rotulo, substring_a_procurar). Conta, por regiao x Bt/Nao
    Bt/Total, quantos registros tem essa substring no campo (comparacao tolerante)."""
    col_map = [(regiao, bt) for regiao in REGIOES_COLUNAS for bt in (None, "SIM", "NAO")]
    resultado = {rotulo: [] for rotulo, _sub in categorias}
    resultado["_total"] = []
    for regiao, bt in col_map:
        regs = _filtra(registros, regiao=regiao, bt=bt)
        resultado["_total"].append(len(regs))
        for rotulo, sub in categorias:
            n = sum(1 for r in regs if sub in _norm(r.get(campo)))
            resultado[rotulo].append(n)
    return col_map, resultado


def _bloco_sim_nao(registros, titulo, campo):
    _, r = _calcular_percentual_campo(registros, campo)
    return {
        "titulo": titulo,
        "linhas": [
            {"label": "Nº Respondido", "valores": r["respondido"]},
            {"label": "COM uso", "valores": r["com"]},
            {"label": "SEM uso", "valores": r["sem"]},
            {"label": "SEM resposta", "valores": r["sem_resposta"]},
            {"label": "% COM uso", "valores": r["pct_com"], "pct": True},
            {"label": "% SEM uso", "valores": r["pct_sem"], "pct": True},
        ],
    }


def _calcular_fbn(registros: list[dict]):
    col_map = [(regiao, bt) for regiao in REGIOES_COLUNAS for bt in (None, "SIM", "NAO")]

    blocos = [
        _bloco_sim_nao(registros, "Utiliza Inoculação de Sementes", "Inoc_Usa"),
        _bloco_sim_nao(registros, "Utiliza Coinoculação", "Coinoc"),
        _bloco_sim_nao(registros, "Utiliza Co + Mo", "CoMo_Usa"),
    ]

    categorias_forma_inoc = [
        ("Industrial", "industrial"), ("Caixa Plantadeira", "caixa"),
        ("Misturador de Sementes", "misturador"), ("Betoneira", "betoneira"),
        ("Lona", "lona"), ("Sulco", "sulco"),
    ]
    _, r_forma = _calcular_categoria_texto(registros, "Inoc_Forma", categorias_forma_inoc)
    blocos.append({
        "titulo": "Forma de Inoculação",
        "linhas": [{"label": nome, "valores": r_forma[nome]} for nome, _ in categorias_forma_inoc],
    })

    categorias_co_mo = [("Via Semente", "semente"), ("Foliar", "foliar")]
    _, r_comomo = _calcular_categoria_texto(registros, "CoMo_Forma", categorias_co_mo)
    blocos.append({
        "titulo": "Forma de Aplicação do Co + Mo",
        "linhas": [{"label": nome, "valores": r_comomo[nome]} for nome, _ in categorias_co_mo],
    })

    return {"col_map": col_map, "blocos": blocos}


def _calcular_tto_sal_cb(registros: list[dict]):
    col_map = [(regiao, bt) for regiao in REGIOES_COLUNAS for bt in (None, "SIM", "NAO")]
    blocos = [
        _bloco_sim_nao(registros, "Tratamento de Sementes", "Tto_Semente"),
        _bloco_sim_nao(registros, "Utilização de SAL + Inseticida", "SAL_CB"),
        _bloco_sim_nao(registros, "Controle Biológico", "Ctrl_Biol"),
    ]
    return {"col_map": col_map, "blocos": blocos}


def _calcular_ins_dess(registros: list[dict]):
    col_map = [(regiao, bt) for regiao in REGIOES_COLUNAS for bt in (None, "SIM", "NAO")]
    _, r = _calcular_percentual_texto(registros, "Dess_Cl", "Inseticida")
    blocos = [{
        "titulo": "Inseticida na Dessecação",
        "linhas": [
            {"label": "COM inseticida na dessecação", "valores": r["com"]},
            {"label": "SEM inseticida na dessecação", "valores": r["sem"]},
            {"label": "% COM inseticida na dessecação", "valores": r["pct_com"], "pct": True},
            {"label": "% SEM inseticida na dessecação", "valores": r["pct_sem"], "pct": True},
        ],
    }]
    return {"col_map": col_map, "blocos": blocos}


def _calcular_categoria_exata(registros: list[dict], campo: str, categorias: list[tuple]):
    """Como _calcular_categoria_texto, mas por IGUALDADE exata (apos normalizar)
    em vez de substring — evita contagem duplicada quando um valor contem o
    texto de outro (ex: 'Folhas largas' e 'Folhas largas e estreitas')."""
    col_map = [(regiao, bt) for regiao in REGIOES_COLUNAS for bt in (None, "SIM", "NAO")]
    resultado = {rotulo: [] for rotulo, _valor in categorias}
    resultado["_total"] = []
    for regiao, bt in col_map:
        regs = _filtra(registros, regiao=regiao, bt=bt)
        resultado["_total"].append(len(regs))
        for rotulo, valor in categorias:
            n = sum(1 for r in regs if _norm(r.get(campo)) == _norm(valor))
            resultado[rotulo].append(n)
    return col_map, resultado


def _calcular_percentual_preenchido(registros: list[dict], campo: str):
    col_map = [(regiao, bt) for regiao in REGIOES_COLUNAS for bt in (None, "SIM", "NAO")]
    chaves = ["com", "sem", "pct_com", "pct_sem"]
    resultado = {k: [] for k in chaves}
    for regiao, bt in col_map:
        regs = _filtra(registros, regiao=regiao, bt=bt)
        r = _bloco_simples_preenchido(regs, campo)
        for k in chaves:
            resultado[k].append(r[k])
    return col_map, resultado


def _calcular_herbicidas(registros: list[dict]):
    col_map = [(regiao, bt) for regiao in REGIOES_COLUNAS for bt in (None, "SIM", "NAO")]

    categorias = [
        ("Herbicida Dessecante", "Herb_Cl1", "Herb_Alv1"),
        ("Herbicida Pré-emergente", "Herb_Cl2", "Herb_Alv2"),
        ("Herbicida Pós-emergente", "Herb_Cl3", "Herb_Alv3"),
        ("Herbicida Pós-emergente Não Seletivo", "Herb_Cl4", "Herb_Alv4"),
    ]

    blocos = []
    for titulo, campo_classe, campo_alvo in categorias:
        _, r = _calcular_percentual_preenchido(registros, campo_classe)
        blocos.append({
            "titulo": titulo,
            "linhas": [
                {"label": "COM aplicação", "valores": r["com"]},
                {"label": "SEM aplicação", "valores": r["sem"]},
                {"label": "% COM aplicação", "valores": r["pct_com"], "pct": True},
                {"label": "% SEM aplicação", "valores": r["pct_sem"], "pct": True},
            ],
        })
        # quebra Folha Larga / Folha Estreita (guardada no proprio campo "alvo")
        categorias_folha = [
            ("Folhas largas", "Folhas largas"),
            ("Folhas estreitas", "Folhas estreitas"),
            ("Folhas largas e estreitas", "Folhas largas e estreitas"),
        ]
        _, r_folha = _calcular_categoria_exata(registros, campo_alvo, categorias_folha)
        blocos.append({
            "titulo": f"{titulo} — Folha Larga / Estreita",
            "linhas": [{"label": nome, "valores": r_folha[nome]} for nome, _ in categorias_folha],
        })

    return {"col_map": col_map, "blocos": blocos}


def _calcular_contagem_por_blocos(registros: list[dict], blocos: list[tuple]):
    """
    Uma linha por registro (questionario), com 3 colunas por bloco de alvo
    passado em 'blocos': Aplicou?, Nº Aplicações, DAE 1ª Aplicação. Versao
    condensada do que seriam as abas Contagem_Pragas/Contagem_Doenças
    originais (534/similar colunas cruas, uma por item individual x pulverizacao).
    """
    headers_id = ["N° Produtor", "Região", "Mesorregião", "Município", "Bt"]
    headers_alvo = []
    for nome, _lista in blocos:
        headers_alvo += [f"{nome} - Aplicou?", f"{nome} - Nº Aplicações", f"{nome} - DAE 1ª Aplicação"]

    linhas = []
    for reg in registros:
        linha = [
            reg.get("Numero_Produtor"),
            reg.get("Regiao"),
            reg.get("Meso_IDR"),
            reg.get("Municipio"),
            reg.get("Bt"),
        ]
        for _nome, lista in blocos:
            daes = _aplicacoes_alvo(reg, lista)
            linha.append("SIM" if daes else "NAO")
            linha.append(len(daes))
            linha.append(min(daes) if daes else None)
        linhas.append(linha)

    return headers_id + headers_alvo, linhas


def _calcular_contagem_pragas(registros: list[dict]):
    return _calcular_contagem_por_blocos(registros, _BLOCOS_ALVO_PRAGAS)


def _calcular_contagem_doencas(registros: list[dict]):
    return _calcular_contagem_por_blocos(registros, _BLOCOS_ALVO_DOENCAS)


def _build_medias_geral(wb: Workbook, registros: list[dict]):
    ws = wb.create_sheet("Médias_Geral")
    col_map, linhas = _calcular_medias_geral(registros)

    # ---- cabecalho ----
    ws.cell(row=1, column=1, value="Item")
    ws.cell(row=1, column=2, value="N° Questionários com respostas")
    col = 3
    for i in range(0, len(col_map), 3):
        regiao, _bt = col_map[i]
        nome = regiao if regiao else "PARANÁ"
        ws.cell(row=1, column=col, value=nome)
        ws.cell(row=2, column=col, value="Total")
        ws.cell(row=2, column=col + 1, value="Cultivares_Bt")
        ws.cell(row=2, column=col + 2, value="Cultivares_Não Bt")
        col += 3

    # ---- linhas ----
    r = 3
    for label, valores in linhas:
        if label is None:
            r += 1
            continue
        ws.cell(row=r, column=1, value=label)
        for i, v in enumerate(valores):
            if v is not None:
                ws.cell(row=r, column=3 + i, value=v)
        # coluna B = mesmo valor da PARANÁ/Total (primeira posicao do col_map)
        if valores[0] is not None:
            ws.cell(row=r, column=2, value=valores[0])
        r += 1


def _bloco_simples_texto(registros, campo, contem):
    """COM = registros cujo campo contem o texto 'contem' (case/acento
    insensiveis); usado p.ex. para 'classe do produto contem Inseticida'."""
    com = sum(1 for r in registros if contem.lower() in _norm(r.get(campo)))
    total = len(registros)
    sem = total - com
    return {
        "com": com, "sem": sem,
        "pct_com": (com / total) if total else None,
        "pct_sem": (sem / total) if total else None,
    }


def _bloco_simples_preenchido(registros, campo):
    """COM = registros em que o campo (texto do alvo) esta preenchido."""
    com = sum(1 for r in registros if r.get(campo))
    total = len(registros)
    sem = total - com
    return {
        "com": com, "sem": sem,
        "pct_com": (com / total) if total else None,
        "pct_sem": (sem / total) if total else None,
    }


# ---------------------------------------------------------------------------
# Funcao principal
# ---------------------------------------------------------------------------
def _escrever_cabecalho_regional(ws, row=1):
    """Escreve o cabecalho padrao: coluna 1 = 'Item', e 24 colunas de dados
    (PARANA + 7 mesorregioes, cada uma com Total/Cultivares_Bt/Cultivares_Nao Bt)."""
    ws.cell(row=row, column=1, value="Item")
    col = 2
    for regiao in REGIOES_COLUNAS:
        nome = regiao if regiao else "PARANÁ"
        ws.cell(row=row, column=col, value=nome)
        ws.cell(row=row + 1, column=col, value="Total")
        ws.cell(row=row + 1, column=col + 1, value="Cultivares_Bt")
        ws.cell(row=row + 1, column=col + 2, value="Cultivares_Não Bt")
        col += 3


def _escrever_linhas_regional(ws, linhas, row_inicio=3, col_dados_inicio=2):
    """linhas: lista de (label, [24 valores]) ou (None, None) para linha em branco."""
    r = row_inicio
    for label, valores in linhas:
        if label is None:
            r += 1
            continue
        ws.cell(row=r, column=1, value=label)
        if valores:
            for i, v in enumerate(valores):
                if v is not None:
                    ws.cell(row=r, column=col_dados_inicio + i, value=v)
        r += 1


def _linhas_distribuicao(resultado):
    """Converte o resultado de _calcular_distribuicao_alvo() em linhas
    (label, valores) prontas para _escrever_linhas_regional()."""
    linhas = []
    for b in resultado["aplicacoes"]["buckets"]:
        linhas.append((f"Nº Levantamentos — {b} aplicações", resultado["aplicacoes"]["contagem"][b]))
    for b in resultado["aplicacoes"]["buckets"]:
        linhas.append((f"% Imóveis — {b} aplicações", resultado["aplicacoes"]["percentual"][b]))
    linhas.append(("Média Aplicações (entre quem aplicou)", resultado["aplicacoes"]["media_aplicantes"]))
    linhas.append(("Média Aplicações (todos os questionários)", resultado["aplicacoes"]["media_total"]))
    linhas.append((None, None))
    for b in resultado["dae"]["buckets"]:
        linhas.append((f"Nº Levantamentos — DAE {b}", resultado["dae"]["contagem"][b]))
    for b in resultado["dae"]["buckets"]:
        linhas.append((f"% Imóveis — DAE {b}", resultado["dae"]["percentual"][b]))
    linhas.append(("Média DAE (entre quem aplicou)", resultado["dae"]["media_aplicantes"]))
    linhas.append(("Média DAE (todos os questionários)", resultado["dae"]["media_total"]))
    return linhas


def _build_distribuicao_sheet(wb: Workbook, nome_aba: str, registros: list[dict], lista_alvo: list[str]):
    ws = wb.create_sheet(nome_aba)
    resultado = _calcular_distribuicao_alvo(registros, lista_alvo)
    _escrever_cabecalho_regional(ws)
    _escrever_linhas_regional(ws, _linhas_distribuicao(resultado))


def _build_blocos_sheet(wb: Workbook, nome_aba: str, resultado: dict):
    ws = wb.create_sheet(nome_aba)
    _escrever_cabecalho_regional(ws)
    linhas = []
    for bloco in resultado["blocos"]:
        linhas.append((bloco["titulo"].upper(), None))
        for linha in bloco["linhas"]:
            linhas.append((linha["label"], linha["valores"]))
        linhas.append((None, None))
    _escrever_linhas_regional(ws, linhas)


def _build_tabela_larga_sheet(wb: Workbook, nome_aba: str, headers: list, linhas: list):
    ws = wb.create_sheet(nome_aba)
    for c, h in enumerate(headers, start=1):
        ws.cell(row=1, column=c, value=h)
    for r, linha in enumerate(linhas, start=2):
        for c, v in enumerate(linha, start=1):
            if v is not None:
                ws.cell(row=r, column=c, value=v)


def _build_grafico_mip_sheet(wb: Workbook, registros: list[dict]):
    ws = wb.create_sheet("Tabelas_Gráficos_MIP")
    dados = _calcular_grafico_mip(registros)
    i_pr = 0  # coluna "PARANÁ - Total", usada para os graficos impressos no Excel

    ws.cell(row=1, column=1, value="Grupo")
    ws.cell(row=1, column=2, value="Nº médio de aplicações (PARANÁ)")
    n = len(dados["grupos_labels"])
    for i, nome in enumerate(dados["grupos_labels"]):
        ws.cell(row=2 + i, column=1, value=nome)
        ws.cell(row=2 + i, column=2, value=dados["media_aplicantes"][i][i_pr])
    chart1 = BarChart()
    chart1.title = "Nº médio de aplicações por grupo (PARANÁ)"
    chart1.add_data(Reference(ws, min_col=2, min_row=1, max_row=1 + n), titles_from_data=True)
    chart1.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1 + n))
    ws.add_chart(chart1, "D2")

    row0 = 3 + n + 2
    ws.cell(row=row0, column=1, value="Espécie de Lagarta")
    ws.cell(row=row0, column=2, value="Nº Questionários (PARANÁ)")
    especies_lag = list(dados["especies_lagarta"].items())
    for i, (nome, valores) in enumerate(especies_lag):
        ws.cell(row=row0 + 1 + i, column=1, value=nome)
        ws.cell(row=row0 + 1 + i, column=2, value=valores[i_pr])
    chart2 = PieChart()
    chart2.title = "Composição de espécies de Lagarta (PARANÁ)"
    chart2.add_data(Reference(ws, min_col=2, min_row=row0, max_row=row0 + len(especies_lag)), titles_from_data=True)
    chart2.set_categories(Reference(ws, min_col=1, min_row=row0 + 1, max_row=row0 + len(especies_lag)))
    ws.add_chart(chart2, "D20")

    row1 = row0 + len(especies_lag) + 3
    ws.cell(row=row1, column=1, value="Espécie de Percevejo")
    ws.cell(row=row1, column=2, value="Nº Questionários (PARANÁ)")
    especies_perc = list(dados["especies_percevejo"].items())
    for i, (nome, valores) in enumerate(especies_perc):
        ws.cell(row=row1 + 1 + i, column=1, value=nome)
        ws.cell(row=row1 + 1 + i, column=2, value=valores[i_pr])
    chart3 = PieChart()
    chart3.title = "Composição de espécies de Percevejo (PARANÁ)"
    chart3.add_data(Reference(ws, min_col=2, min_row=row1, max_row=row1 + len(especies_perc)), titles_from_data=True)
    chart3.set_categories(Reference(ws, min_col=1, min_row=row1 + 1, max_row=row1 + len(especies_perc)))
    ws.add_chart(chart3, "D38")

    row2 = row1 + len(especies_perc) + 3
    buckets = dados["hist_lagartas_aplicacoes"]["buckets"]
    ws.cell(row=row2, column=1, value="Nº Aplicações (Lagartas)")
    ws.cell(row=row2, column=2, value="Nº Questionários (PARANÁ)")
    for i, b in enumerate(buckets):
        ws.cell(row=row2 + 1 + i, column=1, value=b)
        ws.cell(row=row2 + 1 + i, column=2, value=dados["hist_lagartas_aplicacoes"]["contagem"][b][i_pr])
    chart4 = BarChart()
    chart4.title = "Frequência do nº de aplicações — Lagartas (PARANÁ)"
    chart4.add_data(Reference(ws, min_col=2, min_row=row2, max_row=row2 + len(buckets)), titles_from_data=True)
    chart4.set_categories(Reference(ws, min_col=1, min_row=row2 + 1, max_row=row2 + len(buckets)))
    ws.add_chart(chart4, "D56")

    row3 = row2 + len(buckets) + 3
    buckets_dae = dados["hist_lagartas_dae"]["buckets"]
    ws.cell(row=row3, column=1, value="Faixa de DAE (Lagartas)")
    ws.cell(row=row3, column=2, value="Nº Questionários (PARANÁ)")
    for i, b in enumerate(buckets_dae):
        ws.cell(row=row3 + 1 + i, column=1, value=b)
        ws.cell(row=row3 + 1 + i, column=2, value=dados["hist_lagartas_dae"]["contagem"][b][i_pr])
    chart5 = BarChart()
    chart5.title = "Faixa de DAE da 1ª aplicação — Lagartas (PARANÁ)"
    chart5.add_data(Reference(ws, min_col=2, min_row=row3, max_row=row3 + len(buckets_dae)), titles_from_data=True)
    chart5.set_categories(Reference(ws, min_col=1, min_row=row3 + 1, max_row=row3 + len(buckets_dae)))
    ws.add_chart(chart5, "D74")


def _build_grafico_mid_sheet(wb: Workbook, registros: list[dict]):
    ws = wb.create_sheet("Tabelas_Gráficos_MID")
    dados = _calcular_grafico_mid(registros)
    i_pr = 0

    ws.cell(row=1, column=1, value="Grupo")
    ws.cell(row=1, column=2, value="Nº médio de aplicações (PARANÁ)")
    n = len(dados["grupos_labels"])
    for i, nome in enumerate(dados["grupos_labels"]):
        ws.cell(row=2 + i, column=1, value=nome)
        ws.cell(row=2 + i, column=2, value=dados["media_aplicantes"][i][i_pr])
    chart1 = BarChart()
    chart1.title = "Nº médio de aplicações por grupo (PARANÁ)"
    chart1.add_data(Reference(ws, min_col=2, min_row=1, max_row=1 + n), titles_from_data=True)
    chart1.set_categories(Reference(ws, min_col=1, min_row=2, max_row=1 + n))
    ws.add_chart(chart1, "D2")

    row0 = 3 + n + 2
    ws.cell(row=row0, column=1, value="Doença")
    ws.cell(row=row0, column=2, value="Nº Questionários (PARANÁ)")
    composicao = list(dados["composicao_doencas"].items())
    for i, (nome, valores) in enumerate(composicao):
        ws.cell(row=row0 + 1 + i, column=1, value=nome)
        ws.cell(row=row0 + 1 + i, column=2, value=valores[i_pr])
    chart2 = PieChart()
    chart2.title = "Composição percentual das doenças (PARANÁ)"
    chart2.add_data(Reference(ws, min_col=2, min_row=row0, max_row=row0 + len(composicao)), titles_from_data=True)
    chart2.set_categories(Reference(ws, min_col=1, min_row=row0 + 1, max_row=row0 + len(composicao)))
    ws.add_chart(chart2, "D20")

    row1 = row0 + len(composicao) + 3
    buckets = dados["hist_doencas_aplicacoes"]["buckets"]
    ws.cell(row=row1, column=1, value="Nº Aplicações (Doenças)")
    ws.cell(row=row1, column=2, value="Nº Questionários (PARANÁ)")
    for i, b in enumerate(buckets):
        ws.cell(row=row1 + 1 + i, column=1, value=b)
        ws.cell(row=row1 + 1 + i, column=2, value=dados["hist_doencas_aplicacoes"]["contagem"][b][i_pr])
    chart3 = BarChart()
    chart3.title = "Frequência do nº de aplicações — Doenças (PARANÁ)"
    chart3.add_data(Reference(ws, min_col=2, min_row=row1, max_row=row1 + len(buckets)), titles_from_data=True)
    chart3.set_categories(Reference(ws, min_col=1, min_row=row1 + 1, max_row=row1 + len(buckets)))
    ws.add_chart(chart3, "D38")

    row2 = row1 + len(buckets) + 3
    buckets_dae = dados["hist_doencas_dae"]["buckets"]
    ws.cell(row=row2, column=1, value="Faixa de DAE (Doenças)")
    ws.cell(row=row2, column=2, value="Nº Questionários (PARANÁ)")
    for i, b in enumerate(buckets_dae):
        ws.cell(row=row2 + 1 + i, column=1, value=b)
        ws.cell(row=row2 + 1 + i, column=2, value=dados["hist_doencas_dae"]["contagem"][b][i_pr])
    chart4 = BarChart()
    chart4.title = "Faixa de DAE da 1ª aplicação — Doenças (PARANÁ)"
    chart4.add_data(Reference(ws, min_col=2, min_row=row2, max_row=row2 + len(buckets_dae)), titles_from_data=True)
    chart4.set_categories(Reference(ws, min_col=1, min_row=row2 + 1, max_row=row2 + len(buckets_dae)))
    ws.add_chart(chart4, "D56")


def gerar_excel(registros: list[dict], filepath: str = "MesoIDR_Export.xlsx") -> str:
    wb = Workbook()
    wb.remove(wb.active)

    _build_BD(wb)
    _build_total_pr(wb, registros)
    _build_medias_geral(wb, registros)
    _build_tabela_larga_sheet(wb, "Contagem_Pragas", *_calcular_contagem_pragas(registros))
    _build_grafico_mip_sheet(wb, registros)
    _build_distribuicao_sheet(wb, "Lagartas", registros, _LAGARTAS)
    _build_distribuicao_sheet(wb, "Percevejos", registros, _PERCEVEJOS)
    _build_distribuicao_sheet(wb, "OP +Ácaros", registros, _OUTRAS_PRAGAS + ACAROS)
    _build_blocos_sheet(wb, "FBN", _calcular_fbn(registros))
    _build_blocos_sheet(wb, "Tto Sal CB", _calcular_tto_sal_cb(registros))
    _build_blocos_sheet(wb, "Ins_Dess", _calcular_ins_dess(registros))
    _build_blocos_sheet(wb, "Herbicidas", _calcular_herbicidas(registros))
    _build_distribuicao_sheet(wb, "Doenças", registros, DOENCAS)
    _build_tabela_larga_sheet(wb, "Contagem_Doenças", *_calcular_contagem_doencas(registros))
    _build_grafico_mid_sheet(wb, registros)

    wb.save(filepath)
    return filepath


# ---------------------------------------------------------------------------
# Teste local
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import random
    random.seed(7)

    MESOS = REGIOES_IDR
    MUNICS = {
        "Noroeste": ["Campo Mourao", "Umuarama", "Cianorte"],
        "Norte": ["Londrina", "Maringa", "Cornelio Procopio"],
        "Oeste": ["Cascavel", "Toledo", "Foz do Iguacu"],
        "Sudoeste": ["Pato Branco", "Francisco Beltrao"],
        "Centro Sul": ["Guarapuava", "Irati"],
        "Centro": ["Ponta Grossa", "Castro"],
        "Metropolitana e Litoral": ["Curitiba", "Paranagua"],
    }
    CULTIVARES = [" 50I52 RSF IPRO", " 5400 IPRO", " 5644 IPRO", " 6039 IPRO"]

    def _fake_reg(i):
        meso = random.choice(MESOS)
        munic = random.choice(MUNICS[meso])
        reg: dict[str, Any] = {
            "N": i, "Numero_Produtor": f"{i:04d}",
            "Meso_IDR": meso, "Regiao": meso, "Municipio": munic,
            "Area_Soja": round(random.uniform(50, 900), 1),
            "Cultivar": random.choice(CULTIVARES),
            "Bt": random.choice(["SIM", "NAO"]),
            "Produtividade": round(random.uniform(40, 85), 1),
            "Dt_Plantio": f"2024-10-{random.randint(1,28):02d}",
            "Adversidade": random.choice([None, "Seca", "Granizo"]),
            "Sinistro": random.choice(["SIM", "NAO"]),
            "Conhec_MID": random.choice(["SIM", "NAO"]),
            "Utiliza_MID": random.choice(["SIM", "NAO"]),
            "Conhec_MIP": random.choice(["SIM", "NAO"]),
            "Utiliza_MIP": random.choice(["SIM", "NAO"]),
            "Herb_Cl1": "Herbicida", "Herb_Alv1": "Folhas largas", "Herb_Nap1": 1,
            "Dess_Sim": "SIM", "Dess_Dt": "2024-09-20",
            "Dess_Cl": "Herbicida", "Dess_Alv1": "Folhas largas",
            "Tto_Semente": random.choice(["SIM", "NAO"]),
            "SAL_CB": random.choice(["SIM", "NAO"]),
            "Ctrl_Biol": random.choice(["SIM", "NAO"]),
            "Inoc_Usa": random.choice(["SIM", "NAO"]),
            "Inoc_Forma": "Via semente",
            "Coinoc": random.choice(["SIM", "NAO"]),
            "CoMo_Usa": random.choice(["SIM", "NAO"]),
            "CoMo_Forma": random.choice(["Via semente", "Foliar", None]),
        }
        for n in range(1, N_PULV + 1):
            reg[f"P{n}_DAE"] = random.randint(15, 90)
            reg[f"P{n}_Data"] = f"2024-{random.randint(10,12):02d}-{random.randint(1,28):02d}"
            reg[f"P{n}_Cl1"] = "Inseticida"
            reg[f"P{n}_Alv1"] = random.choice(PRAGAS)
        return reg

    registros = [_fake_reg(i) for i in range(1, 21)]
    out = gerar_excel(registros, "/home/claude/out/MesoIDR_Export_teste.xlsx")
    print(f"Gerado: {out} ({len(registros)} registros)")
